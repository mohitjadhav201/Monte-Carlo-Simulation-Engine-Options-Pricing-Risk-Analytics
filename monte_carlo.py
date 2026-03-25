import streamlit as st
import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
from scipy.stats import norm
import io
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image as XLImage
import warnings
warnings.filterwarnings('ignore')

# ─────────────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────────────
st.set_page_config(
    page_title="KAP Analytics | Monte Carlo Simulation",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────
#  CUSTOM CSS
# ─────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap');

html, body, [class*="css"] { font-family: 'Inter', sans-serif; }

/* Dark gradient background */
.stApp {
    background: linear-gradient(135deg, #0a0e1a 0%, #0d1530 40%, #0a1628 100%);
}

/* Header */
.kap-header {
    background: linear-gradient(90deg, #0d47a1 0%, #1565c0 50%, #0d47a1 100%);
    padding: 1.5rem 2.5rem;
    border-radius: 16px;
    margin-bottom: 2rem;
    border: 1px solid rgba(100,181,246,0.3);
    box-shadow: 0 8px 32px rgba(13,71,161,0.4);
    display: flex;
    align-items: center;
    justify-content: space-between;
}
.kap-title { font-size: 2.1rem; font-weight: 800; color: #ffffff; letter-spacing: -0.5px; }
.kap-subtitle { font-size: 0.95rem; color: rgba(255,255,255,0.75); font-weight: 400; margin-top:3px; }
.kap-badge {
    background: rgba(255,255,255,0.15);
    border: 1px solid rgba(255,255,255,0.3);
    border-radius: 20px;
    padding: 6px 16px;
    font-size: 0.75rem;
    color: #fff;
    font-weight: 600;
    letter-spacing: 1px;
}

/* Metric cards */
.metric-card {
    background: linear-gradient(135deg, rgba(13,71,161,0.25) 0%, rgba(21,101,192,0.15) 100%);
    border: 1px solid rgba(100,181,246,0.25);
    border-radius: 14px;
    padding: 1.4rem 1.6rem;
    text-align: center;
    transition: transform 0.2s, box-shadow 0.2s;
    backdrop-filter: blur(10px);
}
.metric-card:hover { transform: translateY(-3px); box-shadow: 0 12px 30px rgba(13,71,161,0.35); }
.metric-label { font-size: 0.78rem; color: rgba(255,255,255,0.55); text-transform: uppercase;
    letter-spacing: 1.2px; font-weight: 600; margin-bottom: 8px; }
.metric-value { font-size: 1.9rem; font-weight: 800; color: #64b5f6; line-height: 1; }
.metric-sub { font-size: 0.75rem; color: rgba(255,255,255,0.4); margin-top: 5px; }

/* Section headers */
.section-header {
    font-size: 1.05rem; font-weight: 700; color: #90caf9;
    border-left: 4px solid #1976d2; padding-left: 12px;
    margin: 1.5rem 0 1rem 0; letter-spacing: 0.3px;
}

/* Sidebar */
[data-testid="stSidebar"] {
    background: linear-gradient(180deg, #0a1628 0%, #0d1f3c 100%);
    border-right: 1px solid rgba(100,181,246,0.15);
}
[data-testid="stSidebar"] .stSlider > div > div { background: #1565c0; }
[data-testid="stSidebar"] label { color: #90caf9 !important; font-weight: 500; font-size:0.85rem; }

/* Number inputs */
[data-testid="stNumberInput"] input { 
    background: rgba(13,71,161,0.2); 
    border: 1px solid rgba(100,181,246,0.3); 
    color: #e3f2fd;
    border-radius: 8px;
}

/* Buttons */
.stButton > button {
    background: linear-gradient(90deg, #1565c0, #0d47a1);
    color: white; border: none; border-radius: 10px;
    font-weight: 600; padding: 0.6rem 2rem;
    transition: all 0.2s;
    box-shadow: 0 4px 15px rgba(13,71,161,0.4);
}
.stButton > button:hover {
    background: linear-gradient(90deg, #1976d2, #1565c0);
    box-shadow: 0 6px 20px rgba(13,71,161,0.6);
    transform: translateY(-1px);
}

/* Download button */
.stDownloadButton > button {
    background: linear-gradient(90deg, #2e7d32, #1b5e20);
    color: white; border: none; border-radius: 10px;
    font-weight: 600; padding: 0.6rem 2rem;
    box-shadow: 0 4px 15px rgba(46,125,50,0.4);
}
.stDownloadButton > button:hover {
    background: linear-gradient(90deg, #388e3c, #2e7d32);
    transform: translateY(-1px);
}

/* Expander */
.streamlit-expanderHeader { color: #90caf9 !important; font-weight: 600; }

/* Info box */
.info-box {
    background: rgba(13,71,161,0.15);
    border: 1px solid rgba(100,181,246,0.2);
    border-radius: 10px;
    padding: 1rem 1.2rem;
    font-size: 0.85rem;
    color: rgba(255,255,255,0.75);
    line-height: 1.6;
}

/* Divider */
hr { border-color: rgba(100,181,246,0.15) !important; }

/* Chart containers */
.chart-container {
    background: rgba(13,71,161,0.1);
    border: 1px solid rgba(100,181,246,0.15);
    border-radius: 14px;
    padding: 1rem;
    margin-bottom: 1rem;
}
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────
#  HEADER
# ─────────────────────────────────────────────────────
st.markdown("""
<div class="kap-header">
  <div>
    <div class="kap-title">📊 KAP Analytics</div>
    <div class="kap-subtitle">Monte Carlo Simulation Engine · Options Pricing & Risk Analytics</div>
  </div>
  <div class="kap-badge">MONTE CARLO SIMULATION</div>
</div>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────────────
#  SIDEBAR
# ─────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Parameters")
    st.markdown("---")

    st.markdown("**📈 Market Parameters**")
    S0 = st.number_input("Initial Stock Price (S₀)", value=1458.70, min_value=1.0, step=0.1, format="%.2f")
    K  = S0   # Strike Price is always equal to Initial Stock Price (ATM)
    st.markdown(
        f"<div style='background:rgba(13,71,161,0.25);border:1px solid rgba(100,181,246,0.3);"
        f"border-radius:8px;padding:8px 12px;margin-bottom:8px;'>"
        f"<span style='color:#90caf9;font-size:0.82rem;font-weight:600;'>Strike Price (K)</span><br>"
        f"<span style='color:#64b5f6;font-size:1.1rem;font-weight:700;'>{K:,.2f}</span>"
        f"<span style='color:rgba(255,255,255,0.45);font-size:0.72rem;'> &nbsp;= S₀ (ATM, locked)</span>"
        f"</div>",
        unsafe_allow_html=True,
    )
    T  = st.number_input("Time to Maturity (T, years)", value=2.0, min_value=0.01, step=0.25, format="%.2f")

    st.markdown("---")
    st.markdown("**📉 Rate Parameters**")
    r  = st.number_input("Risk-free Rate (r, %)", value=6.53, min_value=0.0, max_value=50.0, step=0.01, format="%.2f") / 100
    q  = st.number_input("Dividend Yield (q, %)", value=0.46, min_value=0.0, max_value=50.0, step=0.01, format="%.2f") / 100
    sigma = st.number_input("Annual Volatility (σ, %)", value=30.61, min_value=0.1, max_value=200.0, step=0.01, format="%.2f") / 100

    st.markdown("---")
    st.markdown("**🔢 Simulation Settings**")
    N = st.select_slider("Simulations", options=[1000, 5000, 10000, 25000, 50000, 100000], value=10000)
    trading_days = st.number_input("Trading Days / Year", value=248, min_value=100, max_value=365, step=1)
    use_paths = st.checkbox("Generate Price Paths", value=True, help="Also simulate step-by-step paths for visualization")
    n_paths_display = st.slider("Paths to Display", 10, 200, 50) if use_paths else 50
    seed = st.number_input("Random Seed (0 = use 42)", value=42, min_value=0, step=1)

    st.markdown("---")
    run_btn = st.button("▶ Run Simulation", use_container_width=True)

    st.markdown("""
    <div class="info-box">
    <b>Model:</b> Geometric Brownian Motion (GBM)<br>
    <code>S(T) = S₀ · exp((r-q-½σ²)T + σ√T·Z)</code><br><br>
    <b>Pricing:</b> Risk-neutral discounting<br>
    <code>C = e^(-rT) · E[max(S_T - K, 0)]</code>
    </div>
    """, unsafe_allow_html=True)


# ─────────────────────────────────────────────────────
#  SIMULATION ENGINE
# ─────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def run_simulation(S0, K, T, r, q, sigma, N, trading_days, use_paths, seed):
    # ── Exact multi-step GBM engine ───────────────────
    np.random.seed(42 if seed == 0 else int(seed))
    n_simulations = N
    dt      = 1 / trading_days
    n_steps = int(T * trading_days)
    drift     = (r - q - 0.5 * sigma ** 2) * dt
    diffusion = sigma * np.sqrt(dt)
    Z         = np.random.standard_normal((n_simulations, n_steps))
    exponent  = drift + diffusion * Z
    log_S     = np.log(S0) + np.cumsum(exponent, axis=1)
    S         = np.exp(log_S)
    S         = np.hstack((np.full((n_simulations, 1), S0), S))
    final_prices  = S[:, -1]
    ST            = final_prices
    avg_ST        = final_prices.mean()
    std_ST        = final_prices.std()
    prob_ITM      = np.mean(ST > K)
    payoffs       = np.maximum(ST - K, 0)
    mc_price      = np.exp(-r * T) * np.mean(payoffs)

    # Black-Scholes analytical price & Greeks
    d1 = (np.log(S0 / K) + (r - q + 0.5 * sigma**2) * T) / (sigma * np.sqrt(T))
    d2 = d1 - sigma * np.sqrt(T)
    bs_price = (S0 * np.exp(-q * T) * norm.cdf(d1)
                - K  * np.exp(-r * T) * norm.cdf(d2))
    delta = np.exp(-q * T) * norm.cdf(d1)
    gamma = np.exp(-q * T) * norm.pdf(d1) / (S0 * sigma * np.sqrt(T))
    vega  = S0 * np.exp(-q * T) * norm.pdf(d1) * np.sqrt(T) / 100
    theta = (-(S0 * sigma * np.exp(-q * T) * norm.pdf(d1)) / (2 * np.sqrt(T))
             - r * K * np.exp(-r * T) * norm.cdf(d2)
             + q * S0 * np.exp(-q * T) * norm.cdf(d1)) / 365
    rho   = K * T * np.exp(-r * T) * norm.cdf(d2) / 100

    # Paths for Plotly visualisation — shape (n_steps+1, n_display)
    paths = None
    paths_sample = None
    if use_paths:
        n_display    = min(500, n_simulations)
        paths        = S[:n_display, :].T
        n_xl         = min(100, n_display)
        paths_sample = S[:n_xl, :].T   # smaller sample for Excel matplotlib chart

    return {
        "ST": ST, "avg_ST": avg_ST, "std_ST": std_ST,
        "prob_ITM": prob_ITM, "mc_price": mc_price,
        "bs_price": bs_price, "delta": delta, "gamma": gamma,
        "vega": vega, "theta": theta, "rho": rho,
        "payoffs": payoffs, "paths": paths,
        "paths_sample": paths_sample,
        "d1": d1, "d2": d2,
    }


# ─────────────────────────────────────────────────────
#  EXCEL EXPORT
# ─────────────────────────────────────────────────────
def build_excel(S0, K, T, r, q, sigma, N, trading_days, res):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Colour palette ──
    C_DARK   = "0D1530"
    C_BLUE   = "0D47A1"
    C_LBLUE  = "1565C0"
    C_ACCENT = "42A5F5"
    C_WHITE  = "FFFFFF"
    C_LGRAY  = "E3F2FD"
    C_GOLD   = "FFD700"
    C_GREEN  = "00897B"

    thin  = Side(style="thin",   color="BBDEFB")
    thick = Side(style="medium", color="0D47A1")
    def border(t=thin, b=thin, l=thin, r=thin):
        return Border(top=t, bottom=b, left=l, right=r)
    def fill(hex_): return PatternFill("solid", fgColor=hex_)
    def font(bold=False, color=C_WHITE, sz=11, italic=False):
        return Font(name="Arial", bold=bold, color=color, size=sz, italic=italic)
    def align(h="center", v="center", wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    # ────────────────────────────────────────────────
    #  SHEET 1 – Parameters & Results
    # ────────────────────────────────────────────────
    ws1 = wb.create_sheet("Parameters & Results")
    ws1.sheet_view.showGridLines = False

    # Column widths
    for col, w in [("A",4),("B",32),("C",22),("D",4),("E",28),("F",22),("G",4)]:
        ws1.column_dimensions[col].width = w

    # --- TITLE BANNER ---
    ws1.merge_cells("B2:F2")
    ws1["B2"] = "KAP ANALYTICS"
    ws1["B2"].font      = Font(name="Arial", bold=True, color=C_WHITE, size=20)
    ws1["B2"].fill      = fill(C_BLUE)
    ws1["B2"].alignment = align()

    ws1.merge_cells("B3:F3")
    ws1["B3"] = "Monte Carlo Simulation — Option Pricing Engine"
    ws1["B3"].font      = Font(name="Arial", bold=False, color=C_LGRAY, size=11, italic=True)
    ws1["B3"].fill      = fill(C_LBLUE)
    ws1["B3"].alignment = align()

    ws1.row_dimensions[2].height = 36
    ws1.row_dimensions[3].height = 22

    # --- PARAMETERS TABLE ---
    ws1.merge_cells("B5:C5")
    ws1["B5"] = "INPUT PARAMETERS"
    ws1["B5"].font = Font(name="Arial", bold=True, color=C_WHITE, size=12)
    ws1["B5"].fill = fill(C_BLUE)
    ws1["B5"].alignment = align()
    ws1["C5"].fill = fill(C_BLUE)
    ws1.row_dimensions[5].height = 22

    params = [
        ("Initial Stock Price (S₀)",   S0,            "#,##0.00"),
        ("Strike Price (K)",            K,             "#,##0.00"),
        ("Time to Maturity (T)",        T,             "0.00 \"years\""),
        ("Risk-free Rate (r)",          r,             "0.00%"),
        ("Dividend Yield (q)",          q,             "0.00%"),
        ("Annual Volatility (σ)",       sigma,         "0.00%"),
        ("Number of Simulations",       N,             "#,##0"),
        ("Trading Days / Year",         trading_days,  "#,##0"),
    ]

    for i, (label, val, fmt) in enumerate(params):
        row = 6 + i
        ws1.row_dimensions[row].height = 20
        bg = "EBF5FB" if i % 2 == 0 else C_WHITE
        for c in ["B", "C"]:
            ws1[f"{c}{row}"].fill   = PatternFill("solid", fgColor=bg)
            ws1[f"{c}{row}"].border = border()

        ws1[f"B{row}"] = label
        ws1[f"B{row}"].font      = font(color="0D47A1", sz=10)
        ws1[f"B{row}"].alignment = align(h="left")

        ws1[f"C{row}"] = val
        ws1[f"C{row}"].font         = font(bold=True, color="0D47A1", sz=10)
        ws1[f"C{row}"].alignment    = align()
        ws1[f"C{row}"].number_format = fmt

    # --- RESULTS TABLE ---
    ws1.merge_cells("E5:F5")
    ws1["E5"] = "SIMULATION RESULTS"
    ws1["E5"].font = Font(name="Arial", bold=True, color=C_WHITE, size=12)
    ws1["E5"].fill = fill(C_GREEN)
    ws1["E5"].alignment = align()
    ws1["F5"].fill = fill(C_GREEN)

    results = [
        ("Average Final Stock Price",       res["avg_ST"],   "#,##0.00", False),
        ("Std Dev of Final Stock Price",     res["std_ST"],   "#,##0.00", False),
        ("Median Final Stock Price",         np.median(res["ST"]), "#,##0.00", False),
        ("Probability Final Price > K",      res["prob_ITM"], "0.00%",    False),
        ("MC Call Option Price",             res["mc_price"], "#,##0.00", False),
        ("BS  Call Option Price",            res["bs_price"], "#,##0.00", False),
        ("Delta",                            res["delta"],    "0.0000",   False),
        ("Vega  (per 1% σ change)",          res["vega"],     "#,##0.0000", False),
    ]

    for i, (label, val, fmt, _) in enumerate(results):
        row = 6 + i
        bg = "E8F5E9" if i % 2 == 0 else C_WHITE
        for c in ["E", "F"]:
            ws1[f"{c}{row}"].fill   = PatternFill("solid", fgColor=bg)
            ws1[f"{c}{row}"].border = border()

        ws1[f"E{row}"] = label
        ws1[f"E{row}"].font      = font(color="1B5E20", sz=10)
        ws1[f"E{row}"].alignment = align(h="left")

        ws1[f"F{row}"] = val
        ws1[f"F{row}"].font         = font(bold=True, color="1B5E20", sz=10)
        ws1[f"F{row}"].alignment    = align()
        ws1[f"F{row}"].number_format = fmt

    # --- GREEKS TABLE ---
    ws1.merge_cells("B16:C16")
    ws1["B16"] = "OPTION GREEKS (Black-Scholes)"
    ws1["B16"].font      = Font(name="Arial", bold=True, color=C_WHITE, size=11)
    ws1["B16"].fill      = fill("4527A0")
    ws1["B16"].alignment = align()
    ws1["C16"].fill      = fill("4527A0")
    ws1.row_dimensions[16].height = 22

    greeks = [
        ("Delta (Δ)", res["delta"], "0.0000"),
        ("Gamma (Γ)", res["gamma"], "0.000000"),
        ("Vega  (ν)", res["vega"],  "0.0000"),
        ("Theta (Θ)", res["theta"], "0.0000"),
        ("Rho   (ρ)", res["rho"],   "0.0000"),
    ]
    for i, (lbl, val, fmt) in enumerate(greeks):
        row = 17 + i
        bg = "EDE7F6" if i % 2 == 0 else C_WHITE
        ws1.row_dimensions[row].height = 19
        for c in ["B","C"]:
            ws1[f"{c}{row}"].fill   = PatternFill("solid", fgColor=bg)
            ws1[f"{c}{row}"].border = border()
        ws1[f"B{row}"] = lbl
        ws1[f"B{row}"].font      = font(color="4527A0", sz=10)
        ws1[f"B{row}"].alignment = align(h="left")
        ws1[f"C{row}"] = val
        ws1[f"C{row}"].font          = font(bold=True, color="4527A0", sz=10)
        ws1[f"C{row}"].alignment     = align()
        ws1[f"C{row}"].number_format = fmt

    # Footer
    ws1.row_dimensions[24].height = 16
    ws1.merge_cells("B24:F24")
    ws1["B24"] = f"Generated by KAP Analytics Monte Carlo Engine  ·  N = {N:,} simulations  ·  GBM model"
    ws1["B24"].font      = Font(name="Arial", color="90A4AE", size=8, italic=True)
    ws1["B24"].alignment = align(h="center")

    # ────────────────────────────────────────────────
    #  SHEET 2 – Simulations
    # ────────────────────────────────────────────────
    ws2 = wb.create_sheet("Simulations")
    ws2.sheet_view.showGridLines = False

    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 18

    ST = res["ST"]

    # Headers
    for col, header, bg in [("A","Summary Statistic","0D47A1"),
                              ("B","Value",            "0D47A1"),
                              ("C","Final Price",      "1565C0")]:
        ws2[f"{col}1"].value     = header
        ws2[f"{col}1"].font      = Font(name="Arial", bold=True, color=C_WHITE, size=11)
        ws2[f"{col}1"].fill      = fill(bg)
        ws2[f"{col}1"].alignment = align()
        ws2[f"{col}1"].border    = border()
    ws2.row_dimensions[1].height = 22

    # Summary stats in col A/B
    summary_rows = [
        ("Average Final Price",  f"=AVERAGE(C2:C{N+1})"),
        ("Std Deviation",        f"=STDEV(C2:C{N+1})"),
        ("Median Final Price",   f"=MEDIAN(C2:C{N+1})"),
        ("Min Final Price",      f"=MIN(C2:C{N+1})"),
        ("Max Final Price",      f"=MAX(C2:C{N+1})"),
        (f"P(S_T > {K})",        f"=COUNTIF(C2:C{N+1},\">\"&{K})/COUNT(C2:C{N+1})"),
    ]
    for i, (lbl, formula) in enumerate(summary_rows):
        row = 2 + i
        bg = "EBF5FB" if i % 2 == 0 else C_WHITE
        ws2[f"A{row}"] = lbl
        ws2[f"A{row}"].font      = Font(name="Arial", color="0D47A1", size=10)
        ws2[f"A{row}"].fill      = PatternFill("solid", fgColor=bg)
        ws2[f"A{row}"].alignment = align(h="left")
        ws2[f"A{row}"].border    = border()
        ws2[f"B{row}"] = formula
        ws2[f"B{row}"].font         = Font(name="Arial", bold=True, color="0D47A1", size=10)
        ws2[f"B{row}"].fill         = PatternFill("solid", fgColor=bg)
        ws2[f"B{row}"].alignment    = align()
        ws2[f"B{row}"].number_format = "#,##0.00" if i < 5 else "0.00%"
        ws2[f"B{row}"].border       = border()

    # Final prices in col C
    for i, price in enumerate(ST):
        row = 2 + i
        ws2[f"C{row}"] = round(float(price), 6)
        ws2[f"C{row}"].number_format = "#,##0.00"
        if i < 20:  # style only first 20 for speed
            bg = "EBF5FB" if i % 2 == 0 else C_WHITE
            ws2[f"C{row}"].fill   = PatternFill("solid", fgColor=bg)
            ws2[f"C{row}"].border = border()

    # ────────────────────────────────────────────────
    #  SHEET 3 – Distribution Chart Data + Chart
    # ────────────────────────────────────────────────
    ws3 = wb.create_sheet("Distribution Chart")
    ws3.sheet_view.showGridLines = False

    # Histogram bins
    bins = np.linspace(ST.min(), ST.max(), 51)
    counts, edges = np.histogram(ST, bins=bins)
    bin_centers = (edges[:-1] + edges[1:]) / 2

    ws3["A1"] = "Bin Center"
    ws3["B1"] = "Frequency"
    for col in ["A","B"]:
        ws3[f"{col}1"].font      = Font(name="Arial", bold=True, color=C_WHITE, size=11)
        ws3[f"{col}1"].fill      = fill(C_BLUE)
        ws3[f"{col}1"].alignment = align()

    for i, (bc, cnt) in enumerate(zip(bin_centers, counts)):
        ws3[f"A{i+2}"] = round(float(bc), 2)
        ws3[f"B{i+2}"] = int(cnt)
        ws3[f"A{i+2}"].number_format = "#,##0.00"
        ws3.column_dimensions["A"].width = 14
        ws3.column_dimensions["B"].width = 14

    # Bar chart
    chart = BarChart()
    chart.type = "col"
    chart.title = "Distribution of Final Stock Prices"
    chart.y_axis.title = "Frequency"
    chart.x_axis.title = "Final Stock Price"
    chart.width  = 22
    chart.height = 14

    data_ref = Reference(ws3, min_col=2, min_row=1, max_row=len(counts)+1)
    cats_ref = Reference(ws3, min_col=1, min_row=2, max_row=len(counts)+1)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "1976D2"
    chart.series[0].graphicalProperties.line.solidFill = "0D47A1"
    ws3.add_chart(chart, "D2")

    # ────────────────────────────────────────────────
    #  SHEET 4 – Paths data (if available)
    # ────────────────────────────────────────────────
    if res["paths"] is not None:
        ws4 = wb.create_sheet("Sample Paths")
        ws4.sheet_view.showGridLines = False
        paths = res["paths"]  # shape (steps+1, n_display)
        n_show = min(20, paths.shape[1])

        ws4["A1"] = "Step"
        ws4["A1"].font = Font(name="Arial", bold=True, color=C_WHITE, size=11)
        ws4["A1"].fill = fill(C_BLUE)
        ws4["A1"].alignment = align()
        ws4.column_dimensions["A"].width = 8

        for j in range(n_show):
            col_letter = get_column_letter(j + 2)
            ws4[f"{col_letter}1"] = f"Path {j+1}"
            ws4[f"{col_letter}1"].font = Font(name="Arial", bold=True, color=C_WHITE, size=10)
            ws4[f"{col_letter}1"].fill = fill(C_LBLUE)
            ws4[f"{col_letter}1"].alignment = align()
            ws4.column_dimensions[col_letter].width = 12

        for i in range(paths.shape[0]):
            ws4[f"A{i+2}"] = i
            for j in range(n_show):
                ws4[f"{get_column_letter(j+2)}{i+2}"] = round(float(paths[i, j]), 2)

    # ────────────────────────────────────────────────
    #  SHEET 5 – Colourful Simulation Paths PLOT (matplotlib)
    # ────────────────────────────────────────────────
    if res["paths_sample"] is not None:
        ws5 = wb.create_sheet("Simulation Paths Plot")
        ws5.sheet_view.showGridLines = False

        # Title banner
        ws5.merge_cells("B2:N2")
        ws5["B2"] = "Simulated Stock Price Paths  ·  GBM Monte Carlo"
        ws5["B2"].font      = Font(name="Arial", bold=True, color=C_WHITE, size=14)
        ws5["B2"].fill      = fill(C_BLUE)
        ws5["B2"].alignment = align()
        ws5.row_dimensions[2].height = 28

        paths_xl = res["paths_sample"]   # (n_steps+1, n_xl)
        n_xl     = paths_xl.shape[1]
        n_steps  = paths_xl.shape[0]
        x_axis   = np.arange(n_steps)

        # Build a vibrant multicolour figure matching the reference image
        fig_xl, ax = plt.subplots(figsize=(14, 7), facecolor="white")
        color_maps = [
            plt.cm.tab20,
            plt.cm.tab20b,
            plt.cm.tab20c,
            plt.cm.Set1,
            plt.cm.Set2,
        ]
        for j in range(n_xl):
            cmap  = color_maps[(j // 20) % len(color_maps)]
            color = cmap((j % 20) / 20)
            ax.plot(x_axis, paths_xl[:, j], linewidth=0.75, alpha=0.80, color=color)

        ax.set_title(f"Simulated Stock Prices Over {T} Years (GBM)  —  {n_xl} paths shown",
                     fontsize=14, fontweight="bold", pad=12)
        ax.set_xlabel("Time Steps", fontsize=12)
        ax.set_ylabel("Stock Price", fontsize=12)
        ax.tick_params(labelsize=10)
        ax.grid(True, linestyle="--", alpha=0.3, linewidth=0.5)
        fig_xl.tight_layout()

        # Save to buffer
        img_buf = io.BytesIO()
        fig_xl.savefig(img_buf, format="png", dpi=130, bbox_inches="tight")
        img_buf.seek(0)
        plt.close(fig_xl)

        # Embed in worksheet
        xl_img = XLImage(img_buf)
        xl_img.anchor = "B4"
        ws5.add_image(xl_img)

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────
#  PLOTLY CHART HELPERS
# ─────────────────────────────────────────────────────
PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(13,21,48,0.0)",
    plot_bgcolor ="rgba(13,21,48,0.0)",
    font=dict(family="Inter", color="#90CAF9"),
    margin=dict(l=50, r=20, t=50, b=50),
    xaxis=dict(gridcolor="rgba(100,181,246,0.12)", zeroline=False,
               linecolor="rgba(100,181,246,0.3)"),
    yaxis=dict(gridcolor="rgba(100,181,246,0.12)", zeroline=False,
               linecolor="rgba(100,181,246,0.3)"),
)


def histogram_chart(ST, K, avg_ST):
    counts, edges = np.histogram(ST, bins=80)
    centers = (edges[:-1] + edges[1:]) / 2
    colors  = ["#42A5F5" if c > K else "#EF5350" for c in centers]

    fig = go.Figure()
    fig.add_trace(go.Bar(x=centers, y=counts, marker_color=colors,
                         marker_line_width=0, name="Final Price",
                         hovertemplate="Price: %{x:,.0f}<br>Count: %{y:,}<extra></extra>"))
    fig.add_vline(x=K,      line_width=2, line_dash="dash",
                  line_color="#FF7043", annotation_text=f"Strike K={K:,.1f}",
                  annotation_font_color="#FF7043", annotation_position="top right")
    fig.add_vline(x=avg_ST, line_width=2, line_dash="dot",
                  line_color="#A5D6A7", annotation_text=f"Mean={avg_ST:,.0f}",
                  annotation_font_color="#A5D6A7", annotation_position="top left")
    layout = {**PLOTLY_LAYOUT,
              "title": dict(text="Final Stock Price Distribution", x=0.5, font=dict(size=16)),
              "xaxis_title": "Final Stock Price",
              "yaxis_title": "Frequency",
              "showlegend": False}
    fig.update_layout(**layout)
    return fig


def paths_chart(paths, S0, K, n_show):
    if paths is None:
        return None
    n_steps = paths.shape[0]
    t_axis  = np.linspace(0, 1, n_steps)
    fig = go.Figure()
    for j in range(min(n_show, paths.shape[1])):
        fig.add_trace(go.Scatter(
            x=t_axis, y=paths[:, j],
            mode="lines", line=dict(width=0.8, color="rgba(66,165,245,0.35)"),
            showlegend=False,
            hovertemplate="t=%{x:.2f}<br>Price=%{y:,.1f}<extra></extra>"))
    fig.add_hline(y=K,  line_dash="dash", line_color="#FF7043", line_width=1.5,
                  annotation_text=f"K={K:,.1f}", annotation_font_color="#FF7043")
    fig.add_hline(y=S0, line_dash="dot",  line_color="#A5D6A7", line_width=1.5,
                  annotation_text=f"S₀={S0:,.1f}", annotation_font_color="#A5D6A7")
    layout = {**PLOTLY_LAYOUT,
              "title": dict(text="Simulated Price Paths", x=0.5, font=dict(size=16)),
              "xaxis_title": "Normalised Time (0→T)",
              "yaxis_title": "Stock Price"}
    fig.update_layout(**layout)
    return fig


def distribution_curve_chart(ST, K, avg_ST, std_ST):
    x = np.linspace(max(0, avg_ST - 4*std_ST), avg_ST + 4*std_ST, 400)
    y = norm.pdf(x, avg_ST, std_ST)

    fig = go.Figure()
    # ITM fill
    x_itm = x[x > K]; y_itm = norm.pdf(x_itm, avg_ST, std_ST)
    fig.add_trace(go.Scatter(x=x_itm, y=y_itm, fill="tozeroy",
                             fillcolor="rgba(66,165,245,0.25)",
                             line=dict(color="rgba(66,165,245,0)"),
                             name=f"ITM ({np.mean(ST>K)*100:.1f}%)"))
    # OTM fill
    x_otm = x[x <= K]; y_otm = norm.pdf(x_otm, avg_ST, std_ST)
    fig.add_trace(go.Scatter(x=x_otm, y=y_otm, fill="tozeroy",
                             fillcolor="rgba(239,83,80,0.2)",
                             line=dict(color="rgba(239,83,80,0)"),
                             name=f"OTM ({np.mean(ST<=K)*100:.1f}%)"))
    # Curve
    fig.add_trace(go.Scatter(x=x, y=y, mode="lines",
                             line=dict(color="#42A5F5", width=2.5), name="PDF"))
    fig.add_vline(x=K, line_width=2, line_dash="dash", line_color="#FF7043",
                  annotation_text=f"K={K:,.1f}", annotation_font_color="#FF7043")
    layout = {**PLOTLY_LAYOUT,
              "title": dict(text="Probability Density — Final Stock Price", x=0.5, font=dict(size=16)),
              "xaxis_title": "Final Stock Price",
              "yaxis_title": "Probability Density",
              "legend": dict(x=0.72, y=0.95, bgcolor="rgba(0,0,0,0)")}
    fig.update_layout(**layout)
    return fig


def payoff_chart(ST, K, mc_price, r, T):
    sorted_ST = np.sort(ST)
    payoffs   = np.maximum(sorted_ST - K, 0)
    disc      = np.exp(-r * T)

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=sorted_ST, y=payoffs * disc,
                             mode="lines", line=dict(color="#66BB6A", width=2),
                             name="Discounted Payoff",
                             hovertemplate="S_T=%{x:,.0f}<br>Payoff=%{y:,.2f}<extra></extra>"))
    fig.add_hline(y=mc_price, line_dash="dot", line_color="#FFF176", line_width=1.5,
                  annotation_text=f"MC Price={mc_price:,.2f}",
                  annotation_font_color="#FFF176")
    layout = {**PLOTLY_LAYOUT,
              "title": dict(text="Discounted Call Option Payoff Profile", x=0.5, font=dict(size=16)),
              "xaxis_title": "Final Stock Price",
              "yaxis_title": "Discounted Payoff"}
    fig.update_layout(**layout)
    return fig


# ─────────────────────────────────────────────────────
#  MAIN — RUN & RENDER
# ─────────────────────────────────────────────────────
if "result" not in st.session_state:
    st.session_state.result = None

if run_btn or st.session_state.result is None:
    with st.spinner("⚙️  Running Monte Carlo simulation…"):
        st.session_state.result = run_simulation(
            S0, K, T, r, q, sigma, N, trading_days, use_paths, seed)
        st.session_state.params = dict(S0=S0, K=K, T=T, r=r, q=q,
                                       sigma=sigma, N=N, trading_days=trading_days)

res = st.session_state.result
if res is None:
    st.info("Configure parameters in the sidebar and click **▶ Run Simulation**.")
    st.stop()

# ─── METRIC CARDS ────────────────────────────────────
st.markdown('<div class="section-header">Simulation Results</div>', unsafe_allow_html=True)

cols = st.columns(4)
metrics = [
    ("Avg Final Price",      f"{res['avg_ST']:,.2f}",   "GBM terminal mean"),
    ("Std Deviation",        f"{res['std_ST']:,.2f}",   "Price dispersion"),
    ("P(S_T > K)",           f"{res['prob_ITM']*100:.2f}%", f"Strike K = {K:,.1f}"),
    ("MC Call Price",        f"{res['mc_price']:,.2f}", "Risk-neutral"),
]
for col, (label, value, sub) in zip(cols, metrics):
    with col:
        st.markdown(f"""
        <div class="metric-card">
          <div class="metric-label">{label}</div>
          <div class="metric-value">{value}</div>
          <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

cols2 = st.columns(4)
metrics2 = [
    ("BS Call Price", f"{res['bs_price']:,.2f}", "Analytical"),
    ("Delta (Δ)",     f"{res['delta']:.4f}",     "Rate of price change"),
    ("Gamma (Γ)",     f"{res['gamma']:.6f}",     "Delta sensitivity"),
    ("Vega (ν)",      f"{res['vega']:.4f}",      "Per 1% vol move"),
]
for col, (label, value, sub) in zip(cols2, metrics2):
    with col:
        st.markdown(f"""
        <div class="metric-card">
          <div class="metric-label">{label}</div>
          <div class="metric-value" style="font-size:1.5rem">{value}</div>
          <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)
st.markdown("---")

# ─── CHARTS ──────────────────────────────────────────
st.markdown('<div class="section-header">Visualizations</div>', unsafe_allow_html=True)

tab1, tab2, tab3, tab4 = st.tabs(["📊 Distribution", "🌐 Price Paths",
                                    "📈 Density Curve", "💰 Payoff Profile"])

with tab1:
    st.plotly_chart(histogram_chart(res["ST"], K, res["avg_ST"]),
                    use_container_width=True)

with tab2:
    if res["paths"] is not None:
        fig_p = paths_chart(res["paths"], S0, K, n_paths_display)
        if fig_p:
            st.plotly_chart(fig_p, use_container_width=True)
    else:
        st.info("Enable 'Generate Price Paths' in the sidebar.")

with tab3:
    st.plotly_chart(distribution_curve_chart(res["ST"], K, res["avg_ST"], res["std_ST"]),
                    use_container_width=True)

with tab4:
    st.plotly_chart(payoff_chart(res["ST"], K, res["mc_price"], r, T),
                    use_container_width=True)

st.markdown("---")

# ─── DETAIL TABLES + DOWNLOAD ────────────────────────
col_l, col_r = st.columns([2, 1])

with col_l:
    st.markdown('<div class="section-header">Full Analytics Summary</div>', unsafe_allow_html=True)
    summary_df = pd.DataFrame({
        "Metric": [
            "Initial Stock Price (S₀)", "Strike Price (K)",
            "Time to Maturity (T)", "Risk-free Rate (r)", "Dividend Yield (q)",
            "Annual Volatility (σ)", "Number of Simulations", "Trading Days/Year",
            "─── Results ───",
            "Average Final Price", "Std Deviation", "Median Final Price",
            "P(Final Price > K)", "MC Call Option Price", "BS Call Option Price",
            "─── Greeks ───",
            "Delta (Δ)", "Gamma (Γ)", "Vega (ν)", "Theta (Θ)", "Rho (ρ)",
        ],
        "Value": [
            f"{S0:,.2f}", f"{K:,.2f}", f"{T:.2f} yrs",
            f"{r*100:.2f}%", f"{q*100:.2f}%", f"{sigma*100:.2f}%",
            f"{N:,}", f"{trading_days:,}",
            "─────────────",
            f"{res['avg_ST']:,.4f}", f"{res['std_ST']:,.2f}",
            f"{np.median(res['ST']):,.2f}", f"{res['prob_ITM']*100:.2f}%",
            f"{res['mc_price']:,.4f}", f"{res['bs_price']:,.4f}",
            "─────────────",
            f"{res['delta']:.4f}", f"{res['gamma']:.6f}",
            f"{res['vega']:.4f}", f"{res['theta']:.4f}", f"{res['rho']:.4f}",
        ]
    })
    st.dataframe(summary_df, use_container_width=True, hide_index=True)

with col_r:
    st.markdown('<div class="section-header">Export</div>', unsafe_allow_html=True)
    with st.spinner("Preparing Excel…"):
        excel_bytes = build_excel(S0, K, T, r, q, sigma, N, trading_days, res)

    st.download_button(
        label="⬇️  Download Excel Report",
        data=excel_bytes,
        file_name="KAP_MonteCarlo_OptionPricing.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.markdown("""
    <div class="info-box" style="margin-top:12px">
    <b>Excel includes:</b><br>
    ✅ Parameters & Results<br>
    ✅ All 10,000 Final Prices<br>
    ✅ Summary Statistics (formulas)<br>
    ✅ Distribution Chart<br>
    ✅ Sample Price Paths sheet<br>
    ✅ Colourful Simulation Paths Plot<br>
    ✅ Option Greeks
    </div>""", unsafe_allow_html=True)

# ─── FOOTER ──────────────────────────────────────────
st.markdown("<br>", unsafe_allow_html=True)
st.markdown("""
<div style='text-align:center; color:rgba(255,255,255,0.2); font-size:0.78rem; padding:1rem 0'>
  KAP Analytics · Monte Carlo Simulation Engine · GBM Risk-Neutral Pricing<br>
  For informational purposes only. Not financial advice.
</div>
""", unsafe_allow_html=True)
